"""

Shift Allowance Export Service.
 
Provides functions to export employee shift allowance data

to Excel using openpyxl, with optional filters.

"""
 
# Standard library

from typing import Optional

from datetime import datetime, date
 
# Third-party

from dateutil.relativedelta import relativedelta

from fastapi import HTTPException

from sqlalchemy import func

from sqlalchemy.orm import Session

from openpyxl import Workbook

from openpyxl.utils import get_column_letter
 
# Local imports

from models.models import ShiftAllowances, ShiftMapping, ShiftsAmount
 
 
def _parse_month(month: str, field_name: str) -> date:

    """Convert YYYY-MM string to a date representing the first day of that month."""

    try:

        return datetime.strptime(month, "%Y-%m").date().replace(day=1)

    except ValueError as exc:

        raise HTTPException(

            status_code=400,

            detail=f"{field_name} must be in YYYY-MM format"

        ) from exc
 
 
def _resolve_latest_month(base_query, current_month: date):

    """Return a query filtered to the latest month with data in the last 12 months."""

    for i in range(12):

        check_month = current_month.replace(day=1) - relativedelta(months=i)

        query = base_query.filter(

            func.date_trunc("month", ShiftAllowances.duration_month) == check_month

        )

        if query.first():

            return query

    raise HTTPException(status_code=404, detail="No data found in the last 12 months")
 
 
def _calculate_shift_allowances(db, row, shift_labels, allowance_map):

    """

    Calculate shift entries and total allowance for a single row.
 
    Args:

        db: SQLAlchemy session

        row: ShiftAllowances row

        shift_labels: Dict mapping shift type to label

        allowance_map: Dict mapping shift type to amount
 
    Returns:

        tuple: (shift_entries list, total_allowance float)

    """

    mappings = db.query(ShiftMapping.shift_type, ShiftMapping.days).filter(

        ShiftMapping.shiftallowance_id == row.id

    ).all()
 
    shift_entries = []

    total_allowance = 0.0
 
    for mapping in mappings:

        days = float(mapping.days or 0)

        if days <= 0:

            continue

        shift_type = mapping.shift_type.upper()

        label = shift_labels.get(shift_type, shift_type)

        rate = allowance_map.get(shift_type, 0)

        shift_total = rate * days

        total_allowance += shift_total

        shift_entries.append(

            f"{label}-{int(days)}*{int(rate):,}=₹{int(shift_total):,}"

        )
 
    return shift_entries, total_allowance
 
 
def export_filtered_excel_openpyxl(

    db: Session,

    emp_id: Optional[str] = None,

    account_manager: Optional[str] = None,

    start_month: Optional[str] = None,

    end_month: Optional[str] = None,

    department: Optional[str] = None,

    client: Optional[str] = None,

):

    """

    Export filtered shift allowance records to an openpyxl Workbook.
 
    Optional filters:

        - emp_id

        - account_manager

        - department

        - client

        - start_month / end_month (YYYY-MM)
 
    Returns:

        Workbook: openpyxl workbook object ready for download

    """

    shift_labels = {"A": "A", "B": "B", "C": "C", "PRIME": "PRIME"}
 
    base_query = db.query(

        ShiftAllowances.id,

        ShiftAllowances.emp_id,

        ShiftAllowances.emp_name,

        ShiftAllowances.grade,

        ShiftAllowances.department,

        ShiftAllowances.client,

        ShiftAllowances.project,

        ShiftAllowances.project_code,

        ShiftAllowances.account_manager,

        ShiftAllowances.delivery_manager,

        ShiftAllowances.practice_lead,

        ShiftAllowances.billability_status,

        ShiftAllowances.practice_remarks,

        ShiftAllowances.rmg_comments,

        ShiftAllowances.duration_month,

        ShiftAllowances.payroll_month,

    )
 
    # Apply filters

    if emp_id:

        base_query = base_query.filter(

            func.trim(ShiftAllowances.emp_id) == emp_id.strip()

        )

    if account_manager:

        base_query = base_query.filter(

            func.lower(func.trim(ShiftAllowances.account_manager))

            == account_manager.strip().lower()

        )

    if department:

        base_query = base_query.filter(

            func.lower(func.trim(ShiftAllowances.department))

            == department.strip().lower()

        )

    if client:

        base_query = base_query.filter(

            func.lower(func.trim(ShiftAllowances.client)) == client.strip().lower()

        )
 
    current_month = date.today().replace(day=1)
 
    # Apply month filters

    if start_month or end_month:

        if not start_month:

            raise HTTPException(

                status_code=400,

                detail="start_month is required when end_month is provided"

            )

        start_date = _parse_month(start_month, "start_month")

        end_date = _parse_month(end_month, "end_month") if end_month else start_date

        if start_date > end_date:

            raise HTTPException(

                status_code=400,

                detail="start_month cannot be after end_month"

            )

        query = base_query.filter(

            func.date_trunc("month", ShiftAllowances.duration_month) >= start_date,

            func.date_trunc("month", ShiftAllowances.duration_month) <= end_date,

        )

    else:

        query = _resolve_latest_month(base_query, current_month)
 
    rows = query.all()

    if not rows:

        raise HTTPException(

            status_code=404, detail="No records found for given filters"

        )
 
    allowance_map = {

        item.shift_type.upper(): float(item.amount or 0)

        for item in db.query(ShiftsAmount).all()

    }
 
    # Create workbook

    wb = Workbook()

    ws = wb.active

    ws.title = "Shift Allowances"
 
    headers = [

        "emp_id", "emp_name", "department", "client", "project", "project_code",

        "client_partner", "shift_details", "delivery_manager", "practice_lead",

        "billability_status", "practice_remarks", "rmg_comments",

        "duration_month", "payroll_month", "total_allowance"

    ]

    ws.append(headers)
 
    for row in rows:

        shift_entries, total_allowance = _calculate_shift_allowances(

            db, row, shift_labels, allowance_map

        )

        ws.append([

            row.emp_id,

            row.emp_name,

            row.department,

            row.client,

            row.project,

            row.project_code,

            row.account_manager,

            ", ".join(shift_entries) if shift_entries else None,

            row.delivery_manager,

            row.practice_lead,

            row.billability_status,

            row.practice_remarks,

            row.rmg_comments,

            row.duration_month.strftime("%Y-%m") if row.duration_month else None,

            row.payroll_month.strftime("%Y-%m") if row.payroll_month else None,

            f"₹ {total_allowance:,.2f}",

        ])
 
    # Adjust column widths

    for i, col in enumerate(ws.columns, 1):

        max_length = max(len(str(cell.value or "")) for cell in col)

        ws.column_dimensions[get_column_letter(i)].width = max_length + 2
 
    return wb

 